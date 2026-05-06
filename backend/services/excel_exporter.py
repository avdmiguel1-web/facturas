import os
from pathlib import Path
from typing import Dict, Any, List
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

class ExcelExporter:
    def export_general_data(self, document_id: str, data: Dict[str, Any]) -> str:
        """Export general billing data to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación General"
        
        # Header style
        header_fill = PatternFill(start_color="0033CC", end_color="0033CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Add headers
        headers = ["Campo", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        fields = [
            ("Proveedor", data.get("proveedor", "")),
            ("RIF", data.get("rif", "")),
            ("Fecha", data.get("fecha", "")),
            ("Período", data.get("periodo", "")),
            ("Sub-total (Base Imponible)", data.get("subtotal", "")),
            ("IVA", data.get("iva", "")),
            ("Monto Total", data.get("monto_total", "")),
            ("Moneda", data.get("moneda", "")),
            ("Tipo", data.get("tipo", ""))
        ]
        
        for row, (field, value) in enumerate(fields, 2):
            ws.cell(row=row, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        
        # Save file
        filename = f"facturacion_general_{document_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = EXPORT_DIR / filename
        wb.save(file_path)
        
        return str(file_path)
    
    def export_specialized_data(self, document_id: str, data: Dict[str, Any], assignments: List[Dict[str, Any]]) -> str:
        """Export specialized analysis data to Excel with multiple sheets"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Header style
        header_fill = PatternFill(start_color="0033CC", end_color="0033CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Export Ventas de Terceros
        if data.get('ventas_terceros'):
            ws = wb.create_sheet("Ventas de Terceros")
            headers = ["Descripción", "Cantidad", "Monto en Bs"]
            self._create_table(ws, headers, data['ventas_terceros'], header_fill, header_font)
        
        # Export Telefonía Móvil
        if data.get('telefonia_movil'):
            ws = wb.create_sheet("Telefonía Móvil")
            headers = ["No. Móvil", "Descripción", "Monto en Bs"]
            rows = [[item.get('numero_movil', ''), item.get('descripcion', ''), item.get('monto_bs', '')] 
                   for item in data['telefonia_movil']]
            self._write_table(ws, headers, rows, header_fill, header_font)
        
        # Export Rentas y Servicios
        if data.get('rentas_servicios'):
            ws = wb.create_sheet("Rentas y Servicios")
            headers = ["Descripción", "Fecha Desde", "Fecha Hasta", "Monto en Bs"]
            rows = [[item.get('descripcion', ''), item.get('fecha_desde', ''), 
                    item.get('fecha_hasta', ''), item.get('monto_bs', '')] 
                   for item in data['rentas_servicios']]
            self._write_table(ws, headers, rows, header_fill, header_font)
        
        # Export Resumen de Consumo
        if data.get('resumen_consumo'):
            ws = wb.create_sheet("Resumen de Consumo")
            headers = ["Descripción", "Unidades Consumidas", "Monto en Bs"]
            rows = [[item.get('descripcion', ''), item.get('unidades_consumidas', ''), 
                    item.get('monto_bs', '')] 
                   for item in data['resumen_consumo']]
            self._write_table(ws, headers, rows, header_fill, header_font)
        
        # Export Cost Center Analysis
        if assignments:
            ws = wb.create_sheet("Análisis por Centro de Costo")
            self._create_cost_center_analysis(ws, assignments, header_fill, header_font)
        
        # Save file
        filename = f"analisis_especializado_{document_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = EXPORT_DIR / filename
        wb.save(file_path)
        
        return str(file_path)
    
    def _create_table(self, ws, headers, data, header_fill, header_font):
        """Create a table from dictionary data"""
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, item in enumerate(data, 2):
            for col_idx, key in enumerate(item.keys(), 1):
                ws.cell(row=row_idx, column=col_idx, value=item[key])
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _write_table(self, ws, headers, rows, header_fill, header_font):
        """Write a table with headers and rows"""
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_cost_center_analysis(self, ws, assignments, header_fill, header_font):
        """Create cost center analysis with calculations"""
        headers = ["No. Móvil", "Centro de Costo", "Monto Bs", "Porcentaje"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Calculate total
        total = sum(float(a.get('monto_bs', 0)) for a in assignments)
        
        # Write data
        for row_idx, assignment in enumerate(assignments, 2):
            monto = float(assignment.get('monto_bs', 0))
            percentage = (monto / total * 100) if total > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=assignment.get('numero_movil', ''))
            ws.cell(row=row_idx, column=2, value=assignment.get('cost_center_name', ''))
            ws.cell(row=row_idx, column=3, value=monto)
            ws.cell(row=row_idx, column=4, value=f"{percentage:.2f}%")
        
        # Add totals row
        total_row = len(assignments) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=total).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="100.00%").font = Font(bold=True)
        
        # Group by cost center
        cc_totals = {}
        for a in assignments:
            cc_name = a.get('cost_center_name', 'Sin asignar')
            monto = float(a.get('monto_bs', 0))
            cc_totals[cc_name] = cc_totals.get(cc_name, 0) + monto
        
        # Add cost center summary
        summary_start = total_row + 3
        ws.cell(row=summary_start, column=1, value="Resumen por Centro de Costo").font = Font(bold=True, size=12)
        
        summary_row = summary_start + 2
        ws.cell(row=summary_row, column=1, value="Centro de Costo").fill = header_fill
        ws.cell(row=summary_row, column=1).font = header_font
        ws.cell(row=summary_row, column=2, value="Total Bs").fill = header_fill
        ws.cell(row=summary_row, column=2).font = header_font
        ws.cell(row=summary_row, column=3, value="Porcentaje").fill = header_fill
        ws.cell(row=summary_row, column=3).font = header_font
        
        for idx, (cc_name, cc_total) in enumerate(cc_totals.items(), summary_row + 1):
            percentage = (cc_total / total * 100) if total > 0 else 0
            ws.cell(row=idx, column=1, value=cc_name)
            ws.cell(row=idx, column=2, value=cc_total)
            ws.cell(row=idx, column=3, value=f"{percentage:.2f}%")
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25
